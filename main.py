import json
import time
import random
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from typing import Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# Configuration
from config import DELAY_MIN, DELAY_MAX, BROWSER_TIMEOUT, HEADLESS, VPN_REQUIRE_CONNECTION
from vpn_manager import ExpressVPNManager

def parse_credentials(file_path: str) -> List[tuple]:
    """
    Parse credentials from file.
    Format: email:password entries separated by spaces
    Returns list of (email, password) tuples
    """
    credentials = []
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read().strip()
            if not content:
                print("Warning: credentials.txt is empty")
                return credentials
            
            # Split by spaces and process each entry
            entries = content.split()
            for entry in entries:
                entry = entry.strip()
                if not entry:
                    continue
                
                if ':' not in entry:
                    print(f"Warning: Skipping malformed entry: {entry}")
                    continue
                
                parts = entry.split(':', 1)  # Split only on first colon
                if len(parts) == 2:
                    email, password = parts
                    credentials.append((email.strip(), password.strip()))
                else:
                    print(f"Warning: Skipping malformed entry: {entry}")
    
    except FileNotFoundError:
        print(f"Error: {file_path} not found")
    except Exception as e:
        print(f"Error reading credentials file: {e}")
    
    return credentials

def detect_status(page) -> tuple:
    """
    Detect the login status and extract user information if successful.
    Returns: (status, username, karma, error_message)
    Status can be: 'success', 'invalid', 'banned', 'error'
    """
    try:
        # Wait for page to stabilize
        page.wait_for_load_state('networkidle', timeout=5000)
        time.sleep(2)  # Additional wait for dynamic content
        
        current_url = page.url
        
        # Check for banned/suspended indicators
        page_content = page.content().lower()
        
        # Check for suspension/banned messages
        if 'suspended' in page_content or 'banned' in page_content:
            # Try to extract the specific message
            try:
                error_elem = page.query_selector('div[role="alert"], .error, .message')
                if error_elem:
                    error_text = error_elem.inner_text()
                    if 'suspended' in error_text.lower() or 'banned' in error_text.lower():
                        return ('banned', None, None, error_text)
            except:
                pass
            return ('banned', None, None, 'Account appears to be suspended or banned')
        
        # Check if still on login page (invalid credentials)
        if 'login' in current_url.lower():
            # Look for error messages
            try:
                error_selectors = [
                    'div[role="alert"]',
                    '.AnimatedForm__errorMessage',
                    '[data-testid="login-error"]',
                    '.error',
                    'div:has-text("incorrect")',
                    'div:has-text("wrong")',
                    'div:has-text("invalid")'
                ]
                
                for selector in error_selectors:
                    try:
                        error_elem = page.query_selector(selector)
                        if error_elem:
                            error_text = error_elem.inner_text().lower()
                            if any(word in error_text for word in ['incorrect', 'wrong', 'invalid', 'password', 'username']):
                                return ('invalid', None, None, error_elem.inner_text())
                    except:
                        continue
                
                # If still on login page but no error found, likely invalid
                return ('invalid', None, None, 'Login failed - incorrect credentials')
            except:
                return ('invalid', None, None, 'Login failed - incorrect credentials')
        
        # Check if successfully logged in (redirected away from login page)
        if 'login' not in current_url.lower():
            # Try to extract username and karma
            username, karma = extract_user_info(page)
            if username:
                return ('success', username, karma, None)
            else:
                # Logged in but couldn't extract info
                return ('success', 'Unknown', None, 'Logged in but could not extract user info')
        
        return ('error', None, None, 'Unable to determine login status')
    
    except PlaywrightTimeoutError:
        return ('error', None, None, 'Timeout waiting for page to load')
    except Exception as e:
        return ('error', None, None, f'Error detecting status: {str(e)}')

def extract_user_info(page) -> tuple:
    """
    Extract username and karma from the Reddit page.
    Returns: (username, karma)
    """
    username = None
    karma = None
    
    try:
        # Method 1: Try to get username from profile dropdown/menu
        username_selectors = [
            'button[aria-label*="User"]',
            'a[href*="/user/"]',
            '[data-testid="user-menu"]',
            'button:has-text("/u/")',
            'a[href^="/user/"]'
        ]
        
        for selector in username_selectors:
            try:
                elem = page.query_selector(selector)
                if elem:
                    text = elem.inner_text()
                    # Extract username from text like "/u/username" or "username"
                    if '/u/' in text:
                        username = text.split('/u/')[-1].split()[0].strip()
                    elif text.startswith('u/'):
                        username = text.split('u/')[-1].split()[0].strip()
                    else:
                        username = text.strip()
                    if username:
                        break
            except:
                continue
        
        # Method 2: Navigate to profile page to get username and karma
        if not username:
            try:
                # Try clicking on user menu to get profile link
                user_menu = page.query_selector('button[aria-label*="User"], [data-testid="user-menu"]')
                if user_menu:
                    user_menu.click()
                    time.sleep(1)
                    
                    profile_link = page.query_selector('a[href*="/user/"]')
                    if profile_link:
                        profile_url = profile_link.get_attribute('href')
                        if profile_url:
                            if not profile_url.startswith('http'):
                                profile_url = 'https://www.reddit.com' + profile_url
                            page.goto(profile_url, timeout=10000)
                            time.sleep(2)
                            
                            # Extract username from URL
                            if '/user/' in page.url:
                                username = page.url.split('/user/')[-1].split('/')[0]
            except:
                pass
        
        # If still no username, try to get from current URL
        if not username:
            current_url = page.url
            if '/user/' in current_url:
                username = current_url.split('/user/')[-1].split('/')[0]
        
        # Extract karma from profile page
        if username:
            try:
                # Navigate to user profile if not already there
                if '/user/' not in page.url:
                    page.goto(f'https://www.reddit.com/user/{username}', timeout=10000)
                    time.sleep(2)
                
                # Look for karma elements
                karma_selectors = [
                    '[data-testid="karma"]',
                    '.karma',
                    'span:has-text("karma")',
                    'div:has-text("karma")'
                ]
                
                for selector in karma_selectors:
                    try:
                        karma_elem = page.query_selector(selector)
                        if karma_elem:
                            karma_text = karma_elem.inner_text()
                            # Extract number from text like "1,234 karma" or "1.2k karma"
                            import re
                            numbers = re.findall(r'[\d,\.]+[km]?', karma_text)
                            if numbers:
                                karma = numbers[0]
                                break
                    except:
                        continue
                
                # Alternative: Look for karma in sidebar or profile stats
                if not karma:
                    try:
                        # Look for any element containing karma numbers
                        all_text = page.content()
                        karma_match = re.search(r'(\d{1,3}(?:,\d{3})*(?:\.\d+)?[km]?)\s*karma', all_text, re.IGNORECASE)
                        if karma_match:
                            karma = karma_match.group(1)
                    except:
                        pass
            except:
                pass
        
    except Exception as e:
        print(f"Warning: Error extracting user info: {e}")
    
    return (username, karma)

def login_to_reddit(email: str, password: str, playwright) -> Dict:
    """
    Attempt to login to Reddit with given credentials.
    Returns dictionary with status information.
    """
    result = {
        "email": email,
        "password": password,
        "status": "error",
        "username": None,
        "karma": None,
        "error_message": None
    }
    
    browser = None
    context = None
    page = None
    
    try:
        # Launch browser in incognito mode
        browser = playwright.chromium.launch(
            headless=HEADLESS,
            args=['--incognito']
        )
        
        # Create incognito context
        context = browser.new_context(
            viewport={'width': 1920, 'height': 1080},
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        )
        
        page = context.new_page()
        
        # Navigate to Reddit login page
        print(f"Attempting login for: {email}")
        page.goto('https://www.reddit.com/login', timeout=BROWSER_TIMEOUT)
        
        # Wait for page to load and form to be visible
        try:
            page.wait_for_load_state('networkidle', timeout=10000)
        except:
            pass
        time.sleep(2)
        
        # Try to wait for login form to appear
        try:
            page.wait_for_selector('input[type="text"], input[name="username"], input[id*="username"]', timeout=5000)
        except:
            pass
        
        # Fill in email/username field - expanded selectors
        email_selectors = [
            'input[name="username"]',
            'input[name="user"]',
            'input[type="text"]',
            'input[id*="username"]',
            'input[id*="user"]',
            'input[placeholder*="username" i]',
            'input[placeholder*="email" i]',
            'input[placeholder*="user" i]',
            'input[autocomplete="username"]',
            'form input[type="text"]:first-of-type',
            '#loginUsername',
            '#username'
        ]
        
        email_filled = False
        for selector in email_selectors:
            try:
                email_field = page.query_selector(selector)
                if email_field and email_field.is_visible():
                    email_field.click()  # Click to focus
                    time.sleep(0.5)
                    email_field.fill(email)
                    email_filled = True
                    break
            except:
                continue
        
        if not email_filled:
            # Last resort: try to find any text input in a form
            try:
                text_inputs = page.query_selector_all('form input[type="text"]')
                if text_inputs:
                    email_field = text_inputs[0]
                    if email_field.is_visible():
                        email_field.click()
                        time.sleep(0.5)
                        email_field.fill(email)
                        email_filled = True
            except:
                pass
        
        if not email_filled:
            result["error_message"] = "Could not find email/username field - Reddit page structure may have changed"
            return result
        
        # Fill in password field
        password_selectors = [
            'input[name="password"]',
            'input[type="password"]',
            'input[id*="password"]'
        ]
        
        password_filled = False
        for selector in password_selectors:
            try:
                password_field = page.query_selector(selector)
                if password_field:
                    password_field.fill(password)
                    password_filled = True
                    break
            except:
                continue
        
        if not password_filled:
            result["error_message"] = "Could not find password field"
            return result
        
        # Submit the form
        time.sleep(1)  # Small delay before submitting
        submit_selectors = [
            'button[type="submit"]',
            'button:has-text("Log In")',
            'button:has-text("Sign In")',
            'input[type="submit"]'
        ]
        
        form_submitted = False
        for selector in submit_selectors:
            try:
                submit_button = page.query_selector(selector)
                if submit_button:
                    submit_button.click()
                    form_submitted = True
                    break
            except:
                continue
        
        if not form_submitted:
            # Try pressing Enter on password field
            try:
                password_field.press('Enter')
                form_submitted = True
            except:
                result["error_message"] = "Could not submit login form"
                return result
        
        # Wait for navigation/response
        try:
            page.wait_for_load_state('networkidle', timeout=15000)
        except:
            pass  # Continue even if timeout
        
        time.sleep(3)  # Wait for any redirects or error messages
        
        # Detect status and extract info
        status, username, karma, error_msg = detect_status(page)
        result["status"] = status
        result["username"] = username
        result["karma"] = karma
        result["error_message"] = error_msg
        
    except PlaywrightTimeoutError:
        result["error_message"] = "Timeout waiting for page to load"
    except Exception as e:
        result["error_message"] = f"Error during login attempt: {str(e)}"
    finally:
        # Clean up
        if page:
            try:
                page.close()
            except:
                pass
        if context:
            try:
                context.close()
            except:
                pass
        if browser:
            try:
                browser.close()
            except:
                pass
    
    return result

def print_result(result: Dict):
    """Print result to console with color coding."""
    status = result["status"]
    email = result["email"]
    
    if status == "success":
        print(f"\n✅ SUCCESS: {email}")
        print(f"   Username: {result['username']}")
        print(f"   Karma: {result['karma'] or 'N/A'}")
    elif status == "invalid":
        print(f"\n❌ INVALID: {email}")
        print(f"   Error: {result['error_message']}")
    elif status == "banned":
        print(f"\n⚠️  BANNED: {email}")
        print(f"   Error: {result['error_message']}")
    else:
        print(f"\n⚠️  ERROR: {email}")
        print(f"   Error: {result['error_message']}")

def save_results(results: List[Dict], file_path: str = "results.json"):
    """Save results to JSON file."""
    try:
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(results, f, indent=2, ensure_ascii=False)
        print(f"\nResults saved to {file_path}")
    except Exception as e:
        print(f"Error saving results: {e}")

def save_results_excel(results: List[Dict], file_path: str = "reddit_login_results.xlsx"):
    """Save results to Excel file with all details."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Reddit Login Results"
        
        # Header row
        headers = ["Email", "Password", "Status", "Username", "Karma", "Error Message"]
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Color coding for status column
        status_colors = {
            "success": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # Light green
            "invalid": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # Light red
            "banned": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # Light yellow
            "error": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # Light red
        }
        
        # Add data rows
        for result in results:
            row = [
                result.get("email", ""),
                result.get("password", ""),
                result.get("status", "").upper() if result.get("status") else "",
                result.get("username", "") or "",
                result.get("karma", "") or "",
                result.get("error_message", "") or ""
            ]
            ws.append(row)
            
            # Apply color to status column (column C)
            status_cell = ws.cell(row=ws.max_row, column=3)
            status = result.get("status", "").lower()
            if status in status_colors:
                status_cell.fill = status_colors[status]
        
        # Auto-adjust column widths
        column_widths = {
            "A": 30,  # Email
            "B": 25,  # Password
            "C": 12,  # Status
            "D": 20,  # Username
            "E": 15,  # Karma
            "F": 50   # Error Message
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Wrap text for error message column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Add summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["Reddit Login Bot - Results Summary"])
        ws_summary.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws_summary.append([])
        
        success_count = sum(1 for r in results if r.get("status") == "success")
        invalid_count = sum(1 for r in results if r.get("status") == "invalid")
        banned_count = sum(1 for r in results if r.get("status") == "banned")
        error_count = sum(1 for r in results if r.get("status") == "error")
        
        ws_summary.append(["Total Processed:", len(results)])
        ws_summary.append(["✅ Success:", success_count])
        ws_summary.append(["❌ Invalid:", invalid_count])
        ws_summary.append(["⚠️  Banned:", banned_count])
        ws_summary.append(["⚠️  Errors:", error_count])
        
        # Style summary sheet
        ws_summary['A1'].font = Font(bold=True, size=14)
        for row in ws_summary.iter_rows(min_row=4, max_row=ws_summary.max_row):
            for cell in row:
                if cell.column == 1:
                    cell.font = Font(bold=True)
        
        # Adjust summary column width
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 10
        
        # Save file
        wb.save(file_path)
        print(f"\nExcel results saved to {file_path}")
        
    except Exception as e:
        print(f"Error saving Excel results: {e}")

def main():
    """Main function to orchestrate the login process."""
    print("=" * 60)
    print("Reddit Login Bot")
    print("=" * 60)
    
    # Parse credentials
    credentials = parse_credentials("credentials.txt")
    
    if not credentials:
        print("No valid credentials found. Exiting.")
        return
    
    print(f"\nFound {len(credentials)} credential(s) to process")
    print(f"Delay between attempts: {DELAY_MIN}-{DELAY_MAX} seconds")
    print(f"Headless mode: {HEADLESS}\n")
    
    results = []
    
    with sync_playwright() as playwright:
        for i, (email, password) in enumerate(credentials, 1):
            print(f"\n[{i}/{len(credentials)}] Processing: {email}")
            
            # VPN Check and Auto-Connect
            vpn = ExpressVPNManager(log_callback=lambda msg: print(f"  {msg}"))
            is_connected, loc = vpn.get_status()
            if not is_connected:
                print(f"🔒 VPN not connected. Attempting auto-connect...")
                success, msg = vpn.connect_random_location()
                if not success:
                    from config import VPN_REQUIRE_CONNECTION
                    if VPN_REQUIRE_CONNECTION:
                        print(f"\n❌ Aborting: VPN Auto-connect failed: {msg}. Connection is mandatory.")
                        return
                    print(f"⚠️ VPN Auto-connect failed: {msg}. Continuing as per config.")
                else:
                    print(f"✅ VPN Auto-connected: {msg}")
            
            # Attempt login
            result = login_to_reddit(email, password, playwright)
            results.append(result)
            
            # Print result
            print_result(result)
            
            # Delay before next attempt (except for last one)
            if i < len(credentials):
                delay = random.uniform(DELAY_MIN, DELAY_MAX)
                print(f"\nWaiting {delay:.1f} seconds before next attempt...")
                time.sleep(delay)
    
    # Save results to both JSON and Excel
    save_results(results)
    save_results_excel(results)
    
    # Print summary
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    success_count = sum(1 for r in results if r["status"] == "success")
    invalid_count = sum(1 for r in results if r["status"] == "invalid")
    banned_count = sum(1 for r in results if r["status"] == "banned")
    error_count = sum(1 for r in results if r["status"] == "error")
    
    print(f"Total processed: {len(results)}")
    print(f"✅ Success: {success_count}")
    print(f"❌ Invalid: {invalid_count}")
    print(f"⚠️  Banned: {banned_count}")
    print(f"⚠️  Errors: {error_count}")
    print("=" * 60)

if __name__ == "__main__":
    main()

