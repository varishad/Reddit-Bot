"""
Utility script to generate Excel report from credentials and existing results.
Can be run independently to create/update the Excel report.
"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def parse_credentials(file_path: str):
    """Parse credentials from file."""
    credentials = []
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read().strip()
            if not content:
                return credentials
            
            entries = content.split()
            for entry in entries:
                entry = entry.strip()
                if not entry:
                    continue
                
                if ':' not in entry:
                    continue
                
                parts = entry.split(':', 1)
                if len(parts) == 2:
                    email, password = parts
                    credentials.append((email.strip(), password.strip()))
    except Exception as e:
        print(f"Error reading credentials: {e}")
    
    return credentials

def load_existing_results(file_path: str = "results.json"):
    """Load existing results from JSON file if it exists."""
    if os.path.exists(file_path):
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"Warning: Could not load existing results: {e}")
    return []

def create_excel_report(credentials, existing_results=None, output_file="reddit_login_results.xlsx"):
    """Create Excel report from credentials and existing results."""
    if existing_results is None:
        existing_results = []
    
    # Create a dictionary of existing results by email for quick lookup
    results_dict = {r.get("email"): r for r in existing_results if r.get("email")}
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reddit Login Results"
    
    # Header row
    headers = ["Email", "Password", "Status", "Username", "Karma", "Error Message"]
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Color coding for status column
    status_colors = {
        "success": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "invalid": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "banned": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "error": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "not processed": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    }
    
    # Add data rows
    for email, password in credentials:
        if email in results_dict:
            # Use existing result
            result = results_dict[email]
            row = [
                result.get("email", email),
                result.get("password", password),
                result.get("status", "").upper() if result.get("status") else "",
                result.get("username", "") or "",
                result.get("karma", "") or "",
                result.get("error_message", "") or ""
            ]
            status = result.get("status", "").lower()
        else:
            # Not processed yet
            row = [
                email,
                password,
                "NOT PROCESSED",
                "",
                "",
                ""
            ]
            status = "not processed"
        
        ws.append(row)
        
        # Apply color to status column
        status_cell = ws.cell(row=ws.max_row, column=3)
        if status in status_colors:
            status_cell.fill = status_colors[status]
    
    # Auto-adjust column widths
    column_widths = {
        "A": 30,  # Email
        "B": 25,  # Password
        "C": 15,  # Status
        "D": 20,  # Username
        "E": 15,  # Karma
        "F": 50   # Error Message
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Wrap text for error message column
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["Reddit Login Bot - Results Summary"])
    ws_summary.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws_summary.append([])
    
    processed_results = [r for r in existing_results if r.get("status")]
    success_count = sum(1 for r in processed_results if r.get("status") == "success")
    invalid_count = sum(1 for r in processed_results if r.get("status") == "invalid")
    banned_count = sum(1 for r in processed_results if r.get("status") == "banned")
    error_count = sum(1 for r in processed_results if r.get("status") == "error")
    not_processed = len(credentials) - len(processed_results)
    
    ws_summary.append(["Total Accounts:", len(credentials)])
    ws_summary.append(["✅ Success:", success_count])
    ws_summary.append(["❌ Invalid:", invalid_count])
    ws_summary.append(["⚠️  Banned:", banned_count])
    ws_summary.append(["⚠️  Errors:", error_count])
    ws_summary.append(["⏳ Not Processed:", not_processed])
    
    # Style summary sheet
    ws_summary['A1'].font = Font(bold=True, size=14)
    for row in ws_summary.iter_rows(min_row=4, max_row=ws_summary.max_row):
        for cell in row:
            if cell.column == 1:
                cell.font = Font(bold=True)
    
    # Adjust summary column width
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 10
    
    # Save file
    wb.save(output_file)
    print(f"\n✅ Excel report generated: {output_file}")
    print(f"   Total accounts: {len(credentials)}")
    print(f"   Processed: {len(processed_results)}")
    print(f"   Not processed: {not_processed}")

if __name__ == "__main__":
    print("Generating Excel report...")
    credentials = parse_credentials("credentials.txt")
    existing_results = load_existing_results()
    create_excel_report(credentials, existing_results)
    print("\nDone!")

