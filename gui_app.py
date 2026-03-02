"""
Reddit Bot Commercial Version - GUI Application
Professional GUI with login, activation, dashboard, and bot interface.
"""
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
import threading
import os
import json
from datetime import datetime
from database import Database
from bot_engine import RedditBotEngine

class LoginWindow:
    def __init__(self, root):
        self.root = root
        self.root.title("Reddit Bot - Login")
        self.root.geometry("450x350")
        self.root.resizable(False, False)
        self.center_window()
        
        # Try to initialize database
        try:
            self.db = Database()
        except Exception as e:
            messagebox.showerror("Database Error", f"Failed to connect to database:\n{str(e)}")
            root.destroy()
            return
        
        # Load saved credentials if exists
        self.saved_credentials_file = "saved_credentials.json"
        self.load_saved_credentials()
        
        self.setup_ui()
    
    def center_window(self):
        """Center the window on screen."""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def setup_ui(self):
        """Setup login UI."""
        # Title
        title_frame = tk.Frame(self.root, bg="#2c3e50", height=80)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)
        
        title_label = tk.Label(
            title_frame, 
            text="Reddit Bot", 
            font=("Arial", 24, "bold"),
            bg="#2c3e50",
            fg="white"
        )
        title_label.pack(pady=20)
        
        # Main frame
        main_frame = tk.Frame(self.root, padx=40, pady=30)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # License Key
        tk.Label(main_frame, text="License Key:", font=("Arial", 10)).grid(row=0, column=0, sticky=tk.W, pady=10)
        self.license_entry = tk.Entry(main_frame, font=("Arial", 11), width=30)
        self.license_entry.grid(row=0, column=1, pady=10, padx=10)
        self.license_entry.bind('<Return>', lambda e: self.password_entry.focus())
        
        # Password
        tk.Label(main_frame, text="Password:", font=("Arial", 10)).grid(row=1, column=0, sticky=tk.W, pady=10)
        self.password_entry = tk.Entry(main_frame, font=("Arial", 11), show="*", width=30)
        self.password_entry.grid(row=1, column=1, pady=10, padx=10)
        self.password_entry.bind('<Return>', lambda e: self.login())
        
        # Auto-login checkbox
        self.remember_var = tk.BooleanVar(value=True)
        remember_check = tk.Checkbutton(
            main_frame,
            text="Remember credentials (auto-login next time)",
            variable=self.remember_var,
            font=("Arial", 9)
        )
        remember_check.grid(row=2, column=0, columnspan=2, pady=5)
        
        # Buttons frame
        button_frame = tk.Frame(main_frame)
        button_frame.grid(row=3, column=0, columnspan=2, pady=20)
        
        login_btn = tk.Button(
            button_frame,
            text="Login",
            command=self.login,
            bg="#3498db",
            fg="white",
            font=("Arial", 11, "bold"),
            width=12,
            height=2,
            cursor="hand2"
        )
        login_btn.pack(side=tk.LEFT, padx=5)
        
        activate_btn = tk.Button(
            button_frame,
            text="Activate Account",
            command=self.show_activation,
            bg="#27ae60",
            fg="white",
            font=("Arial", 12, "bold"),
            width=18,
            height=2,
            cursor="hand2"
        )
        activate_btn.pack(side=tk.LEFT, padx=5)
        
        # Status label
        self.status_label = tk.Label(main_frame, text="", fg="red", font=("Arial", 9))
        self.status_label.grid(row=4, column=0, columnspan=2, pady=10)
        
        # Focus on license entry
        self.license_entry.focus()
        
        # Try auto-login if credentials saved
        if self.saved_license_key and self.saved_password:
            self.auto_login()
    
    def load_saved_credentials(self):
        """Load saved credentials from file."""
        self.saved_license_key = None
        self.saved_password = None
        try:
            if os.path.exists(self.saved_credentials_file):
                with open(self.saved_credentials_file, 'r') as f:
                    data = json.load(f)
                    self.saved_license_key = data.get('license_key')
                    self.saved_password = data.get('password')
        except:
            pass
    
    def save_credentials(self, license_key, password):
        """Save credentials to file for auto-login."""
        try:
            data = {
                'license_key': license_key,
                'password': password
            }
            with open(self.saved_credentials_file, 'w') as f:
                json.dump(data, f)
        except:
            pass
    
    def clear_saved_credentials(self):
        """Clear saved credentials."""
        try:
            if os.path.exists(self.saved_credentials_file):
                os.remove(self.saved_credentials_file)
        except:
            pass
    
    def auto_login(self):
        """Auto-login with saved credentials - skip login window if successful."""
        if self.saved_license_key and self.saved_password:
            # Try to authenticate silently
            try:
                success, error_msg, user_data = self.db.authenticate_user(self.saved_license_key, self.saved_password)
                
                if success:
                    # Check if account is activated
                    if user_data and user_data.get('is_active', False):
                        # Successfully authenticated and activated - go directly to dashboard
                        # Use after() to ensure UI is fully initialized before destroying
                        self.root.after(100, self._open_dashboard)
                        return
                    else:
                        # Account not activated - show login window with activation prompt
                        try:
                            self.license_entry.insert(0, self.saved_license_key)
                            self.password_entry.insert(0, self.saved_password)
                            self.root.after(500, lambda: self.show_activation_for_login(self.saved_license_key))
                        except:
                            # Widget might be destroyed, just show activation
                            self.root.after(500, lambda: self.show_activation_for_login(self.saved_license_key))
                        return
                else:
                    # Authentication failed - clear saved credentials and show login
                    self.clear_saved_credentials()
                    try:
                        self.license_entry.insert(0, self.saved_license_key)
                        self.password_entry.insert(0, self.saved_password)
                    except:
                        pass  # Widget might not exist yet
            except Exception as e:
                # Error during auto-login - show login window
                try:
                    self.license_entry.insert(0, self.saved_license_key)
                    self.password_entry.insert(0, self.saved_password)
                except:
                    pass  # Widget might not exist yet
    
    def _open_dashboard(self):
        """Open dashboard window (called from auto-login)."""
        self.root.destroy()
        dashboard = tk.Tk()
        DashboardWindow(dashboard, self.db)
        dashboard.mainloop()
    
    def login(self):
        """Handle login."""
        license_key = self.license_entry.get().strip().upper()
        password = self.password_entry.get()
        
        if not license_key or not password:
            self.status_label.config(text="Please enter both license key and password")
            return
        
        self.status_label.config(text="Authenticating...", fg="blue")
        self.root.update()
        
        try:
            success, error_msg, user_data = self.db.authenticate_user(license_key, password)
            
            if success:
                # Save credentials if remember is checked
                if self.remember_var.get():
                    self.save_credentials(license_key, password)
                
                # Check if account is activated
                if user_data and not user_data.get('is_active', False):
                    # Account not activated - show activation window
                    self.status_label.config(text="Account not activated. Please activate first.", fg="orange")
                    self.root.update()
                    # Show activation window
                    self.show_activation_for_login(license_key)
                    return
                
                # Account is active - proceed to dashboard
                # Save credentials if remember is checked (already done above, but ensure it's saved)
                if self.remember_var.get():
                    self.save_credentials(license_key, password)
                
                self.root.destroy()
                # Open main dashboard
                dashboard = tk.Tk()
                DashboardWindow(dashboard, self.db)
                dashboard.mainloop()
            else:
                self.status_label.config(text=error_msg or "Login failed", fg="red")
        except Exception as e:
            self.status_label.config(text=f"Error: {str(e)}", fg="red")
    
    def show_activation(self):
        """Show activation window."""
        activation_window = tk.Toplevel(self.root)
        activation_window.title("Activate Account")
        activation_window.geometry("500x250")
        activation_window.resizable(False, False)
        activation_window.transient(self.root)
        activation_window.grab_set()
        
        # Center window
        activation_window.update_idletasks()
        x = (activation_window.winfo_screenwidth() // 2) - (500 // 2)
        y = (activation_window.winfo_screenheight() // 2) - (250 // 2)
        activation_window.geometry(f'500x250+{x}+{y}')
        
        main_frame = tk.Frame(activation_window, padx=40, pady=40)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        tk.Label(main_frame, text="Activation Code:", font=("Arial", 11, "bold")).pack(anchor=tk.W, pady=(0, 10))
        activation_entry = tk.Entry(main_frame, font=("Arial", 12), width=40)
        activation_entry.pack(pady=5, ipady=5)
        activation_entry.focus()
        
        # Info label
        info_label = tk.Label(main_frame, text="Enter the activation code provided to you", 
                             fg="gray", font=("Arial", 9))
        info_label.pack(pady=(0, 10))
        
        status_label = tk.Label(main_frame, text="", fg="red", font=("Arial", 9), wraplength=400)
        status_label.pack(pady=10)
        
        def activate():
            code = activation_entry.get().strip().upper()
            if not code:
                status_label.config(text="Please enter activation code", fg="red")
                return
            
            status_label.config(text="Activating...", fg="blue")
            activation_window.update()
            
            try:
                success, error_msg = self.db.activate_account(code)
                if success:
                    messagebox.showinfo("Success", "Account activated successfully!\nYou can now login.")
                    activation_window.destroy()
                else:
                    status_label.config(text=error_msg or "Activation failed", fg="red")
            except Exception as e:
                status_label.config(text=f"Error: {str(e)}", fg="red")
        
        tk.Button(
            main_frame,
            text="Activate",
            command=activate,
            bg="#27ae60",
            fg="white",
            font=("Arial", 11, "bold"),
            width=20,
            height=2,
            cursor="hand2"
        ).pack(pady=15)
        
        activation_entry.bind('<Return>', lambda e: activate())
    
    def show_activation_for_login(self, license_key):
        """Show activation window after login attempt (when account not activated)."""
        activation_window = tk.Toplevel(self.root)
        activation_window.title("Activate Account - Required")
        activation_window.geometry("500x280")
        activation_window.resizable(False, False)
        activation_window.transient(self.root)
        activation_window.grab_set()
        
        # Center window
        activation_window.update_idletasks()
        x = (activation_window.winfo_screenwidth() // 2) - (500 // 2)
        y = (activation_window.winfo_screenheight() // 2) - (280 // 2)
        activation_window.geometry(f'500x280+{x}+{y}')
        
        main_frame = tk.Frame(activation_window, padx=40, pady=40)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        tk.Label(main_frame, text="Account Activation Required", 
                font=("Arial", 14, "bold"), fg="orange").pack(pady=10)
        
        tk.Label(main_frame, text="Your account is not activated yet.", 
                font=("Arial", 10)).pack(pady=5)
        tk.Label(main_frame, text="Please enter your activation code to continue.", 
                font=("Arial", 10)).pack(pady=(0, 15))
        
        tk.Label(main_frame, text="Activation Code:", font=("Arial", 11, "bold")).pack(anchor=tk.W, pady=(0, 10))
        activation_entry = tk.Entry(main_frame, font=("Arial", 12), width=40)
        activation_entry.pack(pady=5, ipady=5)
        activation_entry.focus()
        
        status_label = tk.Label(main_frame, text="", fg="red", font=("Arial", 9), wraplength=400)
        status_label.pack(pady=10)
        
        def activate():
            code = activation_entry.get().strip().upper()
            if not code:
                status_label.config(text="Please enter activation code", fg="red")
                return
            
            status_label.config(text="Activating...", fg="blue")
            activation_window.update()
            
            try:
                success, error_msg = self.db.activate_account(code)
                if success:
                    messagebox.showinfo("Success", "Account activated successfully!\nYou can now use the application.")
                    activation_window.destroy()
                    # Try login again
                    self.login()
                else:
                    status_label.config(text=error_msg or "Activation failed", fg="red")
            except Exception as e:
                status_label.config(text=f"Error: {str(e)}", fg="red")
        
        tk.Button(
            main_frame,
            text="Activate",
            command=activate,
            bg="#27ae60",
            fg="white",
            font=("Arial", 11, "bold"),
            width=20,
            height=2,
            cursor="hand2"
        ).pack(pady=15)
        
        activation_entry.bind('<Return>', lambda e: activate())


class DashboardWindow:
    def __init__(self, root, db):
        self.root = root
        self.db = db
        self.bot_engine = None
        self.current_session_id = None
        self.is_running = False
        
        self.root.title("Reddit Bot - Dashboard")
        self.root.geometry("1000x700")
        
        # Handle window close - disconnect VPN and close
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # VPN connection status tracking
        self.vpn_connected = False
        self.vpn_manager = None
        self.pre_vpn_connected = None
        self.pre_vpn_location = None
        
        self.setup_ui()
        self.load_stats()
        
        # Start remains enabled; it will auto-connect VPN if needed
        self.start_btn.config(state=tk.NORMAL, bg="#10b981", text="▶️  Start Bot")
        
        # Do not log VPN status on startup; only show logs after Start Bot is clicked
        
        # Update VPN status periodically
        self.vpn_status_update_interval = 30000  # 30 seconds
        self.schedule_vpn_update()
    
    def on_closing(self):
        """Handle window close event - close without logging out."""
        if self.is_running:
            if not messagebox.askyesno("Bot Running", "Bot is currently running. Stop and close application?"):
                return
            self.stop_bot()
        
        # Disconnect ExpressVPN before closing
        try:
            import time
            if self.vpn_manager:
                self.log("Disconnecting ExpressVPN...")
                self.vpn_manager.disconnect()
                time.sleep(1)  # Give it a moment to disconnect
            else:
                # Try to disconnect even if manager not initialized
                from vpn_manager import ExpressVPNManager
                vpn_manager = ExpressVPNManager(log_callback=None)
                if vpn_manager.is_available():
                    is_connected, _ = vpn_manager.get_status()
                    if is_connected:
                        self.log("Disconnecting ExpressVPN...")
                        vpn_manager.disconnect()
        except Exception as e:
            # Don't block closing if VPN disconnect fails
            pass
        
        # End session if exists
        if self.current_session_id:
            try:
                self.db.end_session(self.current_session_id)
            except:
                pass
        
        # Close the window
        self.root.destroy()
    
    def setup_ui(self):
        """Setup modern professional dashboard UI."""
        # Modern color scheme
        bg_color = "#f5f7fa"
        header_color = "#1a1a2e"
        card_bg = "#ffffff"
        primary_color = "#667eea"
        success_color = "#10b981"
        danger_color = "#ef4444"
        text_dark = "#1f2937"
        text_light = "#6b7280"
        
        self.root.configure(bg=bg_color)
        
        # Top header bar with gradient effect
        header_frame = tk.Frame(self.root, bg=header_color, height=70)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        # Title with icon
        title_container = tk.Frame(header_frame, bg=header_color)
        title_container.pack(side=tk.LEFT, padx=25, pady=15)
        
        tk.Label(
            title_container,
            text="🤖 Reddit Bot",
            font=("Segoe UI", 18, "bold"),
            bg=header_color,
            fg="white"
        ).pack(side=tk.LEFT)
        
        tk.Label(
            title_container,
            text="Dashboard",
            font=("Segoe UI", 14),
            bg=header_color,
            fg="#a0aec0"
        ).pack(side=tk.LEFT, padx=(10, 0))
        
        # Logout button (modern style)
        logout_btn = tk.Button(
            header_frame,
            text="🚪 Logout",
            command=self.logout,
            bg=danger_color,
            fg="white",
            font=("Segoe UI", 10, "bold"),
            cursor="hand2",
            relief=tk.FLAT,
            padx=20,
            pady=8,
            bd=0,
            activebackground="#dc2626",
            activeforeground="white"
        )
        logout_btn.pack(side=tk.RIGHT, padx=25, pady=15)
        
        # Rights label in header (right side)
        rights_label = tk.Label(
            header_frame,
            text="© 2025 All rights reserved by TasksACE.",
            font=("Segoe UI", 9),
            bg=header_color,
            fg="#cbd5e1"
        )
        rights_label.pack(side=tk.RIGHT, padx=(0, 10), pady=15)
        
        # Main container with padding (more compact)
        main_container = tk.Frame(self.root, bg=bg_color)
        main_container.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)
        
        # Left panel - Stats and Controls (modern cards)
        left_panel = tk.Frame(main_container, width=360, bg=bg_color)
        left_panel.pack(side=tk.LEFT, fill=tk.BOTH, padx=(0, 12))
        left_panel.pack_propagate(False)
        
        # Statistics Card (modern card design - more compact)
        # VPN info bar above statistics
        vpn_bar = tk.Frame(left_panel, bg="#f3f4f6")
        vpn_bar.pack(fill=tk.X, pady=(0, 6))
        tk.Label(
            vpn_bar,
            text="🔒 VPN Status",
            font=("Segoe UI", 9, "bold"),
            bg="#f3f4f6",
            fg="#111827"
        ).pack(side=tk.LEFT, padx=8, pady=6)
        self.vpn_status_label = tk.Label(
            vpn_bar,
            text="Checking...",
            font=("Segoe UI", 9, "bold"),
            bg="#f3f4f6",
            fg="#111827"
        )
        self.vpn_status_label.pack(side=tk.LEFT, padx=(8, 8))
        # Statistics Card (modern card design - more compact)
        stats_card = tk.Frame(left_panel, bg=card_bg, relief=tk.FLAT, bd=0)
        stats_card.pack(fill=tk.BOTH, expand=True, pady=(0, 8))
        
        # Card header
        stats_header = tk.Frame(stats_card, bg=primary_color, height=35)
        stats_header.pack(fill=tk.X)
        stats_header.pack_propagate(False)
        
        tk.Label(
            stats_header,
            text="📊 Statistics",
            font=("Segoe UI", 11, "bold"),
            bg=primary_color,
            fg="white"
        ).pack(side=tk.LEFT, padx=12, pady=8)
        
        # Stats content (more compact)
        stats_content = tk.Frame(stats_card, bg=card_bg)
        stats_content.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)
        # License key row (copy-only)
        license_row = tk.Frame(stats_content, bg=card_bg)
        license_row.pack(fill=tk.X, pady=(0, 6))
        tk.Label(
            license_row,
            text="🔑 License Key:",
            font=("Segoe UI", 9, "bold"),
            bg=card_bg,
            fg=text_dark
        ).pack(side=tk.LEFT)
        self.license_key_var = tk.StringVar(value="")
        license_entry = tk.Entry(
            license_row,
            textvariable=self.license_key_var,
            font=("Consolas", 10),
            relief=tk.FLAT,
            bd=1,
            state="readonly",
            readonlybackground="#f3f4f6"
        )
        license_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(8, 6))
        def _copy_license():
            try:
                self.root.clipboard_clear()
                self.root.clipboard_append(self.license_key_var.get())
            except:
                pass
        tk.Button(
            license_row,
            text="Copy",
            command=_copy_license,
            font=("Segoe UI", 8, "bold"),
            cursor="hand2",
            bg="#e5e7eb",
            fg=text_dark,
            relief=tk.FLAT,
            padx=8,
            pady=2,
            bd=0,
            activebackground="#d1d5db"
        ).pack(side=tk.LEFT)
        
        self.stats_text = tk.Text(
            stats_content,
            height=10,
            font=("Segoe UI", 9),
            wrap=tk.WORD,
            bg=card_bg,
            fg=text_dark,
            relief=tk.FLAT,
            bd=0,
            padx=5,
            pady=5
        )
        self.stats_text.pack(fill=tk.BOTH, expand=True)
        # Configure colored tags for stats display
        try:
            self.stats_text.tag_configure("key", foreground="#111827", font=("Segoe UI", 10, "bold"))
            self.stats_text.tag_configure("ok", foreground="#059669", font=("Segoe UI", 10, "bold"))
            self.stats_text.tag_configure("bad", foreground="#dc2626", font=("Segoe UI", 10, "bold"))
            self.stats_text.tag_configure("muted", foreground="#6b7280", font=("Segoe UI", 9))
            self.stats_text.tag_configure("metric", foreground="#4f46e5", font=("Segoe UI", 10, "bold"))
            # Disable selection in stats text (view-only)
            for ev in ["<Button-1>", "<B1-Motion>", "<Double-1>", "<Triple-1>", "<Control-a>", "<Control-A>", "<Control-c>", "<Control-C>"]:
                self.stats_text.bind(ev, lambda e: "break")
        except:
            pass
        
        # Bot Controls Card (more compact)
        controls_card = tk.Frame(left_panel, bg=card_bg, relief=tk.FLAT, bd=0)
        controls_card.pack(fill=tk.X)
        
        # Card header
        controls_header = tk.Frame(controls_card, bg=success_color, height=35)
        controls_header.pack(fill=tk.X)
        controls_header.pack_propagate(False)
        
        tk.Label(
            controls_header,
            text="⚙️ Bot Controls",
            font=("Segoe UI", 11, "bold"),
            bg=success_color,
            fg="white"
        ).pack(side=tk.LEFT, padx=12, pady=8)
        
        # Controls content (more compact)
        controls_content = tk.Frame(controls_card, bg=card_bg)
        controls_content.pack(fill=tk.X, padx=12, pady=12)
        
        # File selection (modern input - more compact)
        tk.Label(
            controls_content,
            text="📁 Credentials File",
            font=("Segoe UI", 9, "bold"),
            bg=card_bg,
            fg=text_dark
        ).pack(anchor=tk.W, pady=(0, 6))
        
        file_frame = tk.Frame(controls_content, bg=card_bg)
        file_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.file_path_var = tk.StringVar(value="credentials.txt")
        file_entry = tk.Entry(
            file_frame,
            textvariable=self.file_path_var,
            font=("Segoe UI", 10),
            relief=tk.FLAT,
            bd=1,
            bg="#f9fafb",
            fg=text_dark,
            insertbackground=text_dark
        )
        file_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8), ipady=8)
        
        browse_btn = tk.Button(
            file_frame,
            text="Browse",
            command=self.browse_file,
            font=("Segoe UI", 9, "bold"),
            cursor="hand2",
            bg="#e5e7eb",
            fg=text_dark,
            relief=tk.FLAT,
            padx=15,
            pady=8,
            bd=0,
            activebackground="#d1d5db"
        )
        browse_btn.pack(side=tk.LEFT)
        # Keep reference to enable/disable during run
        self.browse_btn = browse_btn
        
        # Parallel browsers setting (more compact)
        parallel_container = tk.Frame(controls_content, bg=card_bg)
        parallel_container.pack(fill=tk.X, pady=(0, 10))
        
        tk.Label(
            parallel_container,
            text="🌐 Parallel Browsers",
            font=("Segoe UI", 9, "bold"),
            bg=card_bg,
            fg=text_dark
        ).pack(side=tk.LEFT, padx=(0, 10))
        
        self.parallel_browsers_var = tk.IntVar(value=1)
        parallel_spinbox = tk.Spinbox(
            parallel_container,
            from_=1,
            to=5,
            textvariable=self.parallel_browsers_var,
            width=6,
            font=("Segoe UI", 11, "bold"),
            relief=tk.FLAT,
            bd=1,
            bg="#f9fafb",
            fg=text_dark,
            buttonbackground=primary_color
        )
        parallel_spinbox.pack(side=tk.LEFT)
        self.parallel_spinbox = parallel_spinbox
        
        tk.Label(
            parallel_container,
            text="(1-5 max)",
            font=("Segoe UI", 9),
            bg=card_bg,
            fg=text_light
        ).pack(side=tk.LEFT, padx=(10, 0))
        
        # (VPN status moved above statistics)
        
        # Start/Stop button (more compact)
        self.start_btn = tk.Button(
            controls_content,
            text="▶️  Start Bot",
            command=self.toggle_bot,
            bg=success_color,
            fg="white",
            font=("Segoe UI", 11, "bold"),
            width=22,
            height=1,
            cursor="hand2",
            relief=tk.FLAT,
            bd=0,
            activebackground="#059669",
            activeforeground="white",
            pady=8
        )
        self.start_btn.pack(pady=(5, 0))
        
        # Right panel - Logs and Results (modern tabs)
        right_panel = tk.Frame(main_container, bg=bg_color)
        right_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Modern notebook style (more compact)
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TNotebook', background=bg_color, borderwidth=0)
        style.configure('TNotebook.Tab', padding=[12, 6], font=('Segoe UI', 9, 'bold'))
        style.map('TNotebook.Tab', background=[('selected', card_bg), ('!selected', '#e5e7eb')])
        
        notebook = ttk.Notebook(right_panel)
        notebook.pack(fill=tk.BOTH, expand=True)
        
        # Logs tab (modern card with toolbar)
        logs_card = tk.Frame(notebook, bg=card_bg)
        notebook.add(logs_card, text="📋 Logs")
        
        logs_header = tk.Frame(logs_card, bg=primary_color, height=30)
        logs_header.pack(fill=tk.X)
        logs_header.pack_propagate(False)
        
        tk.Label(
            logs_header,
            text="Activity Log",
            font=("Segoe UI", 10, "bold"),
            bg=primary_color,
            fg="white"
        ).pack(side=tk.LEFT, padx=12, pady=6)
        
        # Clear logs button in header
        clear_logs_btn = tk.Button(
            logs_header,
            text="🗑️ Clear",
            command=lambda: self.logs_text.delete(1.0, tk.END),
            font=("Segoe UI", 8, "bold"),
            cursor="hand2",
            bg="white",
            fg=primary_color,
            relief=tk.FLAT,
            padx=8,
            pady=3,
            bd=0,
            activebackground="#f3f4f6"
        )
        clear_logs_btn.pack(side=tk.RIGHT, padx=10, pady=4)
        
        # Log filter buttons
        self.log_filter = tk.StringVar(value="all")
        filters_frame = tk.Frame(logs_header, bg=primary_color)
        filters_frame.pack(side=tk.RIGHT, padx=8)
        for label, key in [("All", "all"), ("Info", "info"), ("Warn", "warning"), ("Error", "error"), ("Success", "success")]:
            btn = tk.Radiobutton(
                filters_frame,
                text=label,
                value=key,
                variable=self.log_filter,
                indicatoron=False,
                font=("Segoe UI", 8, "bold"),
                selectcolor="#ffffff",
                fg=primary_color,
                bg="white",
                activebackground="#f3f4f6",
                padx=6,
                pady=2
            )
            btn.pack(side=tk.LEFT, padx=2, pady=4)
        
        self.logs_text = scrolledtext.ScrolledText(
            logs_card,
            font=("Consolas", 9),
            wrap=tk.WORD,
            bg="#1e1e1e",
            fg="#d4d4d4",
            insertbackground="#d4d4d4",
            relief=tk.FLAT,
            bd=0,
            padx=8,
            pady=8
        )
        self.logs_text.pack(fill=tk.BOTH, expand=True)
        
        # Configure log text tags for syntax highlighting
        self.logs_text.tag_config("success", foreground="#10b981")
        self.logs_text.tag_config("error", foreground="#ef4444")
        self.logs_text.tag_config("warning", foreground="#f59e0b")
        self.logs_text.tag_config("info", foreground="#3b82f6")
        self.logs_text.tag_config("timestamp", foreground="#9ca3af")
        
        # Results tab (modern card with summary and table)
        results_card = tk.Frame(notebook, bg=card_bg)
        notebook.add(results_card, text="📊 Results")
        
        results_header = tk.Frame(results_card, bg=success_color, height=30)
        results_header.pack(fill=tk.X)
        results_header.pack_propagate(False)
        
        header_left = tk.Frame(results_header, bg=success_color)
        header_left.pack(side=tk.LEFT, padx=12, pady=6)
        
        tk.Label(
            header_left,
            text="Processing Results",
            font=("Segoe UI", 10, "bold"),
            bg=success_color,
            fg="white"
        ).pack(side=tk.LEFT)
        
        results_toolbar = tk.Frame(results_header, bg=success_color)
        results_toolbar.pack(side=tk.RIGHT, padx=8, pady=4)
        # VPN rotations badge
        self.vpn_rotations_var = tk.StringVar(value="VPN Rotations: 0")
        self.vpn_rotations_label = tk.Label(
            results_toolbar,
            textvariable=self.vpn_rotations_var,
            font=("Segoe UI", 8, "bold"),
            bg="white",
            fg=success_color,
            padx=8,
            pady=3
        )
        self.vpn_rotations_label.pack(side=tk.LEFT, padx=3)
        
        export_btn = tk.Button(
            results_toolbar,
            text="📥 Export",
            command=self.export_excel,
            font=("Segoe UI", 8, "bold"),
            cursor="hand2",
            bg="white",
            fg=success_color,
            relief=tk.FLAT,
            padx=8,
            pady=3,
            bd=0,
            activebackground="#f3f4f6"
        )
        export_btn.pack(side=tk.LEFT, padx=3)
        
        def copy_selected_with_fallback():
            """Copy selected result, or show message if nothing selected."""
            selection = self.results_tree.selection()
            if not selection:
                # Try to get first visible item if nothing selected
                children = self.results_tree.get_children()
                if children:
                    self.results_tree.selection_set(children[0])
                    self._copy_selected_result()
                else:
                    self.log("⚠️  No results to copy. Select a row first or right-click on a row.")
            else:
                self._copy_selected_result()
        
        copy_btn = tk.Button(
            results_toolbar,
            text="📋 Copy",
            command=copy_selected_with_fallback,
            font=("Segoe UI", 8, "bold"),
            cursor="hand2",
            bg="white",
            fg=success_color,
            relief=tk.FLAT,
            padx=8,
            pady=3,
            bd=0,
            activebackground="#f3f4f6"
        )
        copy_btn.pack(side=tk.LEFT, padx=3)
        
        clear_btn = tk.Button(
            results_toolbar,
            text="🗑️ Clear",
            command=self.clear_results,
            font=("Segoe UI", 8, "bold"),
            cursor="hand2",
            bg="white",
            fg=danger_color,
            relief=tk.FLAT,
            padx=8,
            pady=3,
            bd=0,
            activebackground="#f3f4f6"
        )
        clear_btn.pack(side=tk.LEFT, padx=3)
        
        # Progress bar (indeterminate)
        self.batch_progress = ttk.Progressbar(results_toolbar, mode="indeterminate", length=120)
        self.batch_progress.pack(side=tk.LEFT, padx=6)
        
        # Summary stats cards (compact)
        summary_frame = tk.Frame(results_card, bg=card_bg)
        summary_frame.pack(fill=tk.X, padx=8, pady=8)
        
        # Summary cards
        self.summary_total = self._create_stat_card(summary_frame, "Total", "0", "#667eea")
        self.summary_success = self._create_stat_card(summary_frame, "✅ Success", "0", "#10b981")
        self.summary_invalid = self._create_stat_card(summary_frame, "❌ Invalid", "0", "#ef4444")
        self.summary_banned = self._create_stat_card(summary_frame, "⚠️ Banned", "0", "#f59e0b")
        self.summary_error = self._create_stat_card(summary_frame, "⚠️ Error", "0", "#ef4444")
        
        # Search frame
        search_frame = tk.Frame(results_card, bg=card_bg)
        search_frame.pack(fill=tk.X, padx=8, pady=(0, 8))
        
        tk.Label(
            search_frame,
            text="🔍 Search:",
            font=("Segoe UI", 9),
            bg=card_bg,
            fg=text_dark
        ).pack(side=tk.LEFT, padx=(0, 8))
        
        self.results_search_var = tk.StringVar()
        self.results_search_var.trace('w', self.filter_results_table)
        search_entry = tk.Entry(
            search_frame,
            textvariable=self.results_search_var,
            font=("Segoe UI", 9),
            relief=tk.FLAT,
            bd=1,
            bg="#f9fafb",
            fg=text_dark,
            width=30
        )
        search_entry.pack(side=tk.LEFT, padx=(0, 8), ipady=4)
        
        # Results table (Treeview)
        table_frame = tk.Frame(results_card, bg=card_bg)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))
        
        # Treeview with scrollbars
        columns = ("#", "Email", "Status", "Reason", "Username", "Karma", "IP", "Country", "Location")
        self.results_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15)
        
        # Configure columns
        self.results_tree.heading("#", text="#")
        self.results_tree.heading("Email", text="Email")
        self.results_tree.heading("Status", text="Status")
        self.results_tree.heading("Reason", text="Reason")
        self.results_tree.heading("Username", text="Username")
        self.results_tree.heading("Karma", text="Karma")
        self.results_tree.heading("IP", text="IP Address")
        self.results_tree.heading("Country", text="Country")
        self.results_tree.heading("Location", text="Location")
        
        # Column widths (optimized for better readability)
        self.results_tree.column("#", width=45, anchor=tk.CENTER)
        self.results_tree.column("Email", width=200, anchor=tk.W)
        self.results_tree.column("Status", width=120, anchor=tk.CENTER)
        self.results_tree.column("Reason", width=250, anchor=tk.W)
        self.results_tree.column("Username", width=130, anchor=tk.W)
        self.results_tree.column("Karma", width=90, anchor=tk.CENTER)
        self.results_tree.column("IP", width=130, anchor=tk.W)
        self.results_tree.column("Country", width=110, anchor=tk.W)
        self.results_tree.column("Location", width=160, anchor=tk.W)
        
        # Scrollbars
        v_scrollbar_results = ttk.Scrollbar(table_frame, orient="vertical", command=self.results_tree.yview)
        h_scrollbar_results = ttk.Scrollbar(table_frame, orient="horizontal", command=self.results_tree.xview)
        self.results_tree.configure(yscrollcommand=v_scrollbar_results.set, xscrollcommand=h_scrollbar_results.set)
        
        # Grid layout
        self.results_tree.grid(row=0, column=0, sticky="nsew")
        v_scrollbar_results.grid(row=0, column=1, sticky="ns")
        h_scrollbar_results.grid(row=1, column=0, sticky="ew")
        
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        
        # Configure tags for color coding (improved colors)
        self.results_tree.tag_configure("success", background="#d1fae5", foreground="#065f46")
        self.results_tree.tag_configure("invalid", background="#fee2e2", foreground="#991b1b")
        self.results_tree.tag_configure("banned", background="#fef3c7", foreground="#92400e")
        self.results_tree.tag_configure("error", background="#fee2e2", foreground="#991b1b")
        
        # Enable sorting and row context menu
        self._setup_results_sorting()
        self._setup_results_context_menu()
        
        # Store results for filtering
        self.all_results = []
        
        # Update VPN status
        self.update_vpn_status()
        
        # Footer status bar
        footer_frame = tk.Frame(self.root, bg="#0f172a", height=26)
        footer_frame.pack(fill=tk.X, side=tk.BOTTOM)
        footer_frame.pack_propagate(False)
        self.footer_label = tk.Label(
            footer_frame,
            text="Elapsed: 00:00:00   Checked: 0   ✅ Good: 0   ❌ Wrong: 0   ⚠️ Banned: 0   Unclear: 0   Login Failed: 0",
            font=("Consolas", 9),
            bg="#0f172a",
            fg="#e5e7eb",
            anchor="w"
        )
        self.footer_label.pack(fill=tk.X, padx=10)
        # Responsive footer updates on window resize
        try:
            self.root.bind("<Configure>", lambda e: self._update_footer_bar())
        except:
            pass
    
    def _setup_results_sorting(self):
        """Setup simple column sorting for results table."""
        try:
            for col in self.results_tree["columns"]:
                def sorter(c=col):
                    data = [(self.results_tree.set(k, c), k) for k in self.results_tree.get_children('')]
                    try:
                        data.sort(key=lambda t: (t[0] is None, t[0]))
                    except:
                        data.sort()
                    for index, (_, k) in enumerate(data):
                        self.results_tree.move(k, '', index)
                self.results_tree.heading(col, text=col, command=sorter)
        except:
            pass
    
    def _setup_results_context_menu(self):
        """Setup right-click context menu for results table."""
        self.results_context_menu = tk.Menu(self.root, tearoff=0)
        self.results_context_menu.add_command(label="📋 Copy Row", command=self._copy_selected_result)
        self.results_context_menu.add_command(label="📋 Copy Email", command=lambda: self._copy_selected_result_field("email"))
        self.results_context_menu.add_command(label="📋 Copy Username", command=lambda: self._copy_selected_result_field("username"))
        self.results_context_menu.add_separator()
        self.results_context_menu.add_command(label="📋 Copy All Results", command=self._copy_all_results)
        
        # Bind right-click to show context menu
        self.results_tree.bind("<Button-3>", self._show_results_context_menu)
        # Also bind on Windows/Linux
        self.results_tree.bind("<Button-2>", self._show_results_context_menu)
    
    def _show_results_context_menu(self, event):
        """Show context menu on right-click."""
        try:
            item = self.results_tree.identify_row(event.y)
            if item:
                self.results_tree.selection_set(item)
                self.results_context_menu.post(event.x_root, event.y_root)
        except:
            pass
    
    def _copy_selected_result(self):
        """Copy selected result row as formatted text."""
        try:
            selection = self.results_tree.selection()
            if not selection:
                return
            
            item = self.results_tree.item(selection[0])
            values = item['values']
            
            # Format: Email | Status | Reason | Username | Karma | IP | Country | Location
            formatted = f"""Email: {values[1]}
Status: {values[2]}
Reason: {values[3]}
Username: {values[4]}
Karma: {values[5]}
IP Address: {values[6]}
Country: {values[7]}
Location: {values[8]}"""
            
            self.root.clipboard_clear()
            self.root.clipboard_append(formatted)
            self.log("✅ Copied result to clipboard")
        except Exception as e:
            self.log(f"Error copying result: {str(e)}")
    
    def _copy_selected_result_field(self, field_name):
        """Copy specific field from selected result."""
        try:
            selection = self.results_tree.selection()
            if not selection:
                return
            
            item = self.results_tree.item(selection[0])
            values = item['values']
            
            field_map = {
                "email": 1,
                "username": 4,
                "karma": 5,
                "ip": 6,
                "country": 7,
                "location": 8
            }
            
            if field_name in field_map:
                idx = field_map[field_name]
                value = values[idx] if idx < len(values) else ""
                self.root.clipboard_clear()
                self.root.clipboard_append(str(value))
                self.log(f"✅ Copied {field_name} to clipboard")
        except Exception as e:
            self.log(f"Error copying field: {str(e)}")
    
    def _copy_all_results(self):
        """Copy all results as formatted text."""
        try:
            if not self.all_results:
                return
            
            formatted_lines = []
            formatted_lines.append("=" * 80)
            formatted_lines.append("REDDIT BOT RESULTS")
            formatted_lines.append("=" * 80)
            formatted_lines.append("")
            
            for i, result in enumerate(self.all_results, 1):
                formatted_lines.append(f"Result #{i}")
                formatted_lines.append(f"  Email: {result.get('email', 'N/A')}")
                formatted_lines.append(f"  Status: {result.get('status', 'N/A')}")
                formatted_lines.append(f"  Reason: {result.get('error_message', 'N/A')}")
                formatted_lines.append(f"  Username: {result.get('username', 'N/A')}")
                formatted_lines.append(f"  Karma: {result.get('karma', 'N/A')}")
                formatted_lines.append(f"  IP: {result.get('ip_address', 'N/A')}")
                formatted_lines.append(f"  Country: {result.get('country', 'N/A')}")
                formatted_lines.append(f"  Location: {result.get('location', 'N/A')}")
                formatted_lines.append("")
            
            formatted_text = "\n".join(formatted_lines)
            self.root.clipboard_clear()
            self.root.clipboard_append(formatted_text)
            self.log("✅ Copied all results to clipboard")
        except Exception as e:
            self.log(f"Error copying all results: {str(e)}")
    
    def load_stats(self):
        """Load and display user statistics with modern formatting."""
        try:
            stats = self.db.get_user_stats()
            if stats:
                try:
                    self.stats_text.config(state=tk.NORMAL)
                except:
                    pass
                self.stats_text.delete(1.0, tk.END)
                # Pretty date format: 11 Nov 2025 11:45am
                def _fmt(ts, fallback):
                    try:
                        if not ts:
                            return fallback
                        # Parse as UTC and convert to Bangladesh Time (UTC+6, no DST)
                        dt_utc = datetime.fromisoformat(str(ts).replace('Z', '+00:00'))
                        from datetime import timedelta, timezone
                        bd_tz = timezone(timedelta(hours=6))  # Asia/Dhaka UTC+6
                        dt_bd = dt_utc.astimezone(bd_tz)
                        # Keep month as 'Nov' (Title case); only lowercase am/pm
                        s = dt_bd.strftime("%d %b %Y %I:%M%p")
                        s = s.replace("AM", "am").replace("PM", "pm")
                        return s
                    except:
                        return fallback
                last_login = _fmt(stats.get('last_login'), "Never")
                activated_at = _fmt(stats.get('activated_at'), "Not activated")
                stats_content = f"""🔑 License Key: {stats.get('license_key', 'N/A')}

✅ Status: {'🟢 Active' if stats.get('is_active') else '🔴 Inactive'}

📊 Total Sessions: {stats.get('total_sessions', 0)}

👥 Total Accounts Processed: {stats.get('total_accounts_processed', 0)}

🕐 Last Login: {last_login}

🔐 Activated: {activated_at}
"""
                self.stats_text.insert(1.0, stats_content)
                # Apply color tags line-by-line
                try:
                    content = self.stats_text.get("1.0", tk.END)
                    lines = content.splitlines()
                    idx = 1
                    for line in lines:
                        start = f"{idx}.0"
                        end = f"{idx}.end"
                        if line.startswith("🔑"):
                            self.stats_text.tag_add("key", start, end)
                        elif line.startswith("✅ Status"):
                            if "🟢 Active" in line:
                                self.stats_text.tag_add("ok", start, end)
                            else:
                                self.stats_text.tag_add("bad", start, end)
                        elif line.startswith("📊") or line.startswith("👥"):
                            self.stats_text.tag_add("metric", start, end)
                        elif line.startswith("🕐") or line.startswith("🔐"):
                            self.stats_text.tag_add("muted", start, end)
                        idx += 1
                except:
                    pass
                try:
                    self.stats_text.config(state=tk.DISABLED)
                except:
                    pass
        except Exception as e:
            self.log(f"Error loading stats: {str(e)}")
    
    def update_vpn_status(self):
        """Update VPN status display with current IP and country, and enable/disable Start Bot button."""
        try:
            from vpn_manager import ExpressVPNManager
            from ip_utils import get_ip_info
            
            # Use stored VPN manager if available, otherwise create new one
            if not self.vpn_manager:
                self.vpn_manager = ExpressVPNManager(log_callback=None)
            
            vpn_manager = self.vpn_manager
            if vpn_manager.is_available():
                is_connected, vpn_location = vpn_manager.get_status()
                
                # Update VPN connection status
                self.vpn_connected = is_connected
                
                # Keep Start button enabled; it will auto-connect if needed
                if not self.is_running:
                    self.start_btn.config(
                        state=tk.NORMAL,
                        text="▶️  Start Bot",
                        bg="#10b981",
                        activebackground="#059669"
                    )
                
                # Get IP info
                ip, country, full_location = get_ip_info()
                
                if is_connected:
                    # Build display string
                    if vpn_location and vpn_location != "Connected":
                        display_location = vpn_location if len(vpn_location) < 20 else vpn_location[:17] + "..."
                    else:
                        display_location = "Connected"
                    
                    if ip and country:
                        self.vpn_status_label.config(
                            text=f"🟢 {display_location} | {country} | IP: {ip}",
                            fg="#10b981"
                        )
                    elif ip:
                        self.vpn_status_label.config(
                            text=f"🟢 {display_location} | IP: {ip}",
                            fg="#10b981"
                        )
                    else:
                        self.vpn_status_label.config(
                            text=f"🟢 {display_location}",
                            fg="#10b981"
                        )
                else:
                    # Show friendly message based on error type
                    if "admin" in vpn_location.lower() or "rights" in vpn_location.lower():
                        if ip:
                            self.vpn_status_label.config(
                                text=f"⚠️ Run as admin | IP: {ip}",
                                fg="#f59e0b"
                            )
                        else:
                            self.vpn_status_label.config(
                                text="⚠️ Run as admin for VPN",
                                fg="#f59e0b"
                            )
                    elif "not found" in vpn_location.lower():
                        if ip:
                            self.vpn_status_label.config(
                                text=f"⚠️ ExpressVPN not found | IP: {ip}",
                                fg="#f59e0b"
                            )
                        else:
                            self.vpn_status_label.config(
                                text="⚠️ ExpressVPN not found",
                                fg="#f59e0b"
                            )
                    elif "not connected" in vpn_location.lower():
                        # Show current IP and country even when not connected
                        if ip and country:
                            self.vpn_status_label.config(
                                text=f"🔴 Not Connected | {country} | IP: {ip}",
                                fg="#ef4444"
                            )
                        elif ip:
                            self.vpn_status_label.config(
                                text=f"🔴 Not Connected | IP: {ip}",
                                fg="#ef4444"
                            )
                        else:
                            self.vpn_status_label.config(
                                text="🔴 Not Connected",
                                fg="#ef4444"
                            )
                    else:
                        # Truncate long error messages
                        display_msg = vpn_location if len(vpn_location) < 30 else vpn_location[:27] + "..."
                        if ip:
                            self.vpn_status_label.config(
                                text=f"🔴 {display_msg} | IP: {ip}",
                                fg="#ef4444"
                            )
                        else:
                            self.vpn_status_label.config(
                                text=f"🔴 {display_msg}",
                                fg="#ef4444"
                            )
            else:
                # ExpressVPN not found, but show IP
                self.vpn_connected = False
                # Keep Start enabled even if VPN not found
                if not self.is_running:
                    self.start_btn.config(
                        state=tk.NORMAL,
                        text="▶️  Start Bot",
                        bg="#10b981",
                        activebackground="#059669"
                    )
                
                ip, country, full_location = get_ip_info()
                if ip and country:
                    self.vpn_status_label.config(
                        text=f"⚠️ ExpressVPN not found | {country} | IP: {ip}",
                        fg="#f59e0b"
                    )
                elif ip:
                    self.vpn_status_label.config(
                        text=f"⚠️ ExpressVPN not found | IP: {ip}",
                        fg="#f59e0b"
                    )
                else:
                    self.vpn_status_label.config(
                        text="⚠️ ExpressVPN not found",
                        fg="#f59e0b"
                    )
        except Exception as e:
            # Don't show error details, just indicate check failed
            self.vpn_connected = False
            if not self.is_running:
                self.start_btn.config(
                    state=tk.NORMAL,
                    text="▶️  Start Bot",
                    bg="#10b981",
                    activebackground="#059669"
                )
            self.vpn_status_label.config(
                text="❌ VPN check failed",
                fg="#ef4444"
            )
    
    def schedule_vpn_update(self):
        """Schedule periodic VPN status updates."""
        self.update_vpn_status()
        self.root.after(self.vpn_status_update_interval, self.schedule_vpn_update)
    
    def show_startup_vpn_info(self):
        """Show VPN information when app starts (status only, no auto-connect)."""
        try:
            import time
            from vpn_manager import ExpressVPNManager
            from ip_utils import get_ip_info
            
            # Use root.after to update UI from thread
            def log_message(msg):
                self.root.after(0, lambda: self.log(msg))
            
            vpn_manager = ExpressVPNManager(log_callback=log_message)
            # Store VPN manager for later use (disconnect on close)
            self.root.after(0, lambda: setattr(self, 'vpn_manager', vpn_manager))
            
            # Get IP info before VPN connection
            ip_before, country_before, location_before = get_ip_info()
            
            log_message("=" * 60)
            log_message("🔒 VPN STATUS ON STARTUP")
            log_message("=" * 60)
            
            # Check VPN status (do not auto-connect here)
            if vpn_manager.is_available():
                is_connected, vpn_location = vpn_manager.get_status()
                
                if is_connected:
                    log_message(f"✅ VPN Already Connected: {vpn_location}")
                    self.root.after(100, lambda: setattr(self, 'vpn_connected', True))
                    self.root.after(100, self.update_vpn_status)
                    self.root.after(2000, self.update_vpn_status)  # Double-check after 2 seconds
                else:
                    log_message("⚠️  ExpressVPN is not connected")
                
                # Get IP info after VPN connection (or if already connected)
                time.sleep(2)  # Give VPN time to establish if just connected
                ip, country, location = get_ip_info()
                
                if ip:
                    log_message(f"📍 IP Address: {ip}")
                if country:
                    log_message(f"🌍 Country: {country}")
                if location:
                    log_message(f"📍 Location: {location}")
                
                # Show comparison if we connected
                if ip_before and ip:
                    if ip_before != ip:
                        log_message(f"✅ IP Changed: {ip_before} → {ip} (VPN working!)")
                    else:
                        log_message(f"⚠️  IP Unchanged: {ip} (VPN may still be connecting)")
                
            else:
                log_message("⚠️  ExpressVPN not found")
                if ip_before:
                    log_message(f"📍 IP Address: {ip_before}")
                if country_before:
                    log_message(f"🌍 Country: {country_before}")
                if location_before:
                    log_message(f"📍 Location: {location_before}")
            
            log_message("=" * 60)
            
            # Update VPN status display (no connect attempts)
            self.root.after(500, self.update_vpn_status)
            self.root.after(2000, self.update_vpn_status)
            self.root.after(5000, self.update_vpn_status)
        except Exception as e:
            err_msg = f"Error getting VPN info: {str(e)}"
            self.root.after(0, lambda msg=err_msg: self.log(msg))
    
    def browse_file(self):
        """Browse for credentials file."""
        # Prevent multiple file dialogs
        if hasattr(self, '_browsing') and self._browsing:
            return
        self._browsing = True
        try:
            filename = filedialog.askopenfilename(
                title="Select Credentials File",
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
            )
            if filename:
                self.file_path_var.set(filename)
        finally:
            self._browsing = False
    
    def toggle_bot(self):
        """Start or stop the bot."""
        if not self.is_running:
            self.start_bot()
        else:
            self.stop_bot()
    
    def connect_vpn(self):
        """Manually connect VPN (no auto-connect after stop)."""
        try:
            from vpn_manager import ExpressVPNManager
            from ip_utils import get_ip_info
            if not self.vpn_manager:
                self.vpn_manager = ExpressVPNManager(log_callback=self.log_callback)
            self.log("Connecting VPN (random location)...")
            # Prefer strategic/randomized selection if available
            success = False
            msg = ""
            try:
                from config import VPN_PREFERRED_COUNTRIES, VPN_AVOID_COUNTRIES, VPN_LOCATION_COOLDOWN_SECONDS, VPN_LOCATION_MAX_TRIES_PER_ROTATION
            except Exception:
                VPN_PREFERRED_COUNTRIES, VPN_AVOID_COUNTRIES = [], []
                VPN_LOCATION_COOLDOWN_SECONDS = 900
                VPN_LOCATION_MAX_TRIES_PER_ROTATION = 10
            try:
                # Try strategy (random among preferred/avoid with cooldown)
                success, msg = self.vpn_manager.connect_with_strategy(
                    preferred=VPN_PREFERRED_COUNTRIES,
                    avoid=VPN_AVOID_COUNTRIES,
                    cooldown_seconds=int(VPN_LOCATION_COOLDOWN_SECONDS),
                    max_candidates=int(VPN_LOCATION_MAX_TRIES_PER_ROTATION)
                )
            except:
                success = False
            if not success:
                # Fall back to random location
                success, msg = self.vpn_manager.connect_random_location()
            if not success:
                # Final fallback to smart location
                success, msg = self.vpn_manager.connect(None)
            if success:
                self.vpn_connected = True
                self.update_vpn_status()
                self.log(f"VPN connected: {msg}")
                ip, country, loc = get_ip_info()
                if ip: self.log(f"IP: {ip}")
            else:
                self.log(f"VPN connect failed: {msg}")
        except Exception as e:
            self.log(f"VPN connect error: {str(e)}")
    
    def disconnect_vpn(self):
        """Manually disconnect VPN."""
        try:
            if self.vpn_manager:
                self.log("Disconnecting VPN...")
                self.vpn_manager.disconnect()
                self.vpn_connected = False
                self.update_vpn_status()
        except Exception as e:
            self.log(f"VPN disconnect error: {str(e)}")
    
    def reconnect_vpn(self):
        """Reconnect VPN on demand."""
        try:
            self.disconnect_vpn()
            self.connect_vpn()
        except Exception as e:
            self.log(f"VPN reconnect error: {str(e)}")
    
    def start_bot(self):
        """Start the bot in a separate thread."""
        # Prevent double-clicking and multiple starts
        if self.is_running:
            return  # Already running, ignore click
        
        # Immediately disable button to prevent double-clicks
        self.start_btn.config(state=tk.DISABLED)
        self.root.update_idletasks()  # Force GUI update
        
        try:
            # Auto-connect VPN if not connected
            if not self.vpn_connected:
                try:
                    self.connect_vpn()
                except Exception as e:
                    self.log(f"VPN auto-connect error: {str(e)}")
                self.update_vpn_status()
                if not self.vpn_connected:
                    messagebox.showwarning("VPN", "Could not connect VPN automatically. Please ensure ExpressVPN is logged in.")
                    self.start_btn.config(state=tk.NORMAL, text="▶️  Start Bot", bg="#10b981")
                    return
            
            file_path = self.file_path_var.get()
            if not os.path.exists(file_path):
                messagebox.showerror("Error", f"File not found: {file_path}")
                self.start_btn.config(state=tk.NORMAL, text="▶️  Start Bot", bg="#10b981")
                return
            
            # Initialize footer counters and timer
            import time as _time
            self._footer_counts = {"total": 0, "success": 0, "invalid": 0, "banned": 0, "error": 0}
            self._footer_start_ts = _time.time()
            self._footer_running = True
            self._tick_footer()
            
            self.is_running = True
            # Immediately show Stop to avoid confusion
            self.start_btn.config(text="⏹️  Stop Bot", bg="#ef4444", activebackground="#dc2626", state=tk.NORMAL)
            
            try:
                self.browse_btn.config(state=tk.DISABLED)
            except:
                pass
            try:
                self.parallel_spinbox.config(state="disabled")
            except:
                pass
            # Start progress bar if available
            try:
                if hasattr(self, "batch_progress"):
                    self.batch_progress.start(10)
            except:
                pass
            self.log("=" * 60)
            self.log("Starting Reddit Bot...")
            self.log("=" * 60)
            
            # Create session
            self.current_session_id = self.db.create_session()
            if not self.current_session_id:
                self.log("Error: Failed to create session")
                self.stop_bot()
                return
            
            # Snapshot current VPN state to restore later
            try:
                if self.vpn_manager:
                    is_connected, vpn_location = self.vpn_manager.get_status()
                    self.pre_vpn_connected = bool(is_connected)
                    self.pre_vpn_location = vpn_location if is_connected else None
            except:
                self.pre_vpn_connected = None
                self.pre_vpn_location = None
            
            # Start bot in thread; capture parallel_browsers value on main thread
            try:
                parallel_browsers = int(self.parallel_browsers_var.get())
            except:
                parallel_browsers = 1
            thread = threading.Thread(target=self.run_bot, args=(file_path, parallel_browsers), daemon=True)
            thread.start()
        except Exception as e:
            # If any error occurs, restore button state
            self.log(f"Error starting bot: {str(e)}")
            self.start_btn.config(state=tk.NORMAL, text="▶️  Start Bot", bg="#10b981")
            self.is_running = False

    def _tick_footer(self):
        """Update footer elapsed time every second while running."""
        try:
            if getattr(self, "_footer_running", False):
                import time as _time
                elapsed = int(_time.time() - getattr(self, "_footer_start_ts", _time.time()))
                hh = elapsed // 3600
                mm = (elapsed % 3600) // 60
                ss = elapsed % 60
                c = getattr(self, "_footer_counts", {"total":0,"success":0,"invalid":0,"banned":0,"error":0})
                # 'Unclear' == error; 'Login Failed' == error
                self._last_footer_snapshot = {
                    "elapsed": f"{hh:02d}:{mm:02d}:{ss:02d}",
                    "total": c.get('total',0),
                    "success": c.get('success',0),
                    "invalid": c.get('invalid',0),
                    "banned": c.get('banned',0),
                    "error": c.get('error',0)
                }
                self._update_footer_bar()
                self.root.after(1000, self._tick_footer)
        except:
            pass

    def _update_footer_bar(self):
        """Responsive footer: show full text on wide screens, compact on small screens."""
        try:
            width = self.root.winfo_width()
            snap = getattr(self, "_last_footer_snapshot", None)
            if not snap:
                # Initial/default text
                text_full = "Elapsed: 00:00:00   Checked: 0   ✅ Good: 0   ❌ Wrong: 0   ⚠️ Banned: 0   Unclear: 0   Login Failed: 0"
            else:
                text_full = (
                    f"Elapsed: {snap['elapsed']}   "
                    f"Checked: {snap['total']}   ✅ Good: {snap['success']}   "
                    f"❌ Wrong: {snap['invalid']}   ⚠️ Banned: {snap['banned']}   "
                    f"Unclear: {snap['error']}   Login Failed: {snap['error']}"
                )
            # Compact version for narrow windows
            text_compact = "Elapsed: {e}   ✅ {s}   ❌ {i}   ⚠️ {b}".format(
                e=(snap['elapsed'] if snap else "00:00:00"),
                s=(snap['success'] if snap else 0),
                i=(snap['invalid'] if snap else 0),
                b=(snap['banned'] if snap else 0),
            )
            # Threshold ~900px for full footer
            if width >= 900:
                self.footer_label.config(text=text_full, font=("Consolas", 9))
            elif width >= 700:
                # Mid size: slightly smaller font, semi-compact
                mid = "Elapsed: {e}   Checked: {t}   ✅ {s}   ❌ {i}   ⚠️ {b}".format(
                    e=(snap['elapsed'] if snap else "00:00:00"),
                    t=(snap['total'] if snap else 0),
                    s=(snap['success'] if snap else 0),
                    i=(snap['invalid'] if snap else 0),
                    b=(snap['banned'] if snap else 0),
                )
                self.footer_label.config(text=mid, font=("Consolas", 8))
            else:
                self.footer_label.config(text=text_compact, font=("Consolas", 8))
        except:
            pass

    def progress_callback(self, snapshot):
        """Receive progress updates from engine and refresh footer counts."""
        try:
            self._footer_counts.update({
                "total": snapshot.get("total", self._footer_counts.get("total", 0)),
                "success": snapshot.get("success", self._footer_counts.get("success", 0)),
                "invalid": snapshot.get("invalid", self._footer_counts.get("invalid", 0)),
                "banned": snapshot.get("banned", self._footer_counts.get("banned", 0)),
                "error": snapshot.get("error", self._footer_counts.get("error", 0)),
            })
            # Update VPN rotations badge if provided
            if "vpn_rotations" in snapshot:
                rotations = snapshot.get("vpn_rotations", 0)
                self.vpn_rotations_var.set(f"VPN Rotations: {rotations}")
        except:
            pass
    
    def run_bot(self, file_path, parallel_browsers):
        """Run the bot (called in thread)."""
        try:
            # Pass GUI VPN manager to engine and skip engine-side VPN init to avoid re-connect hangs
            self.bot_engine = RedditBotEngine(
                self.db,
                self.current_session_id,
                self.log_callback,
                external_vpn_manager=self.vpn_manager,
                skip_vpn_init=True,
                progress_callback=lambda snap: self.root.after(0, lambda s=snap: self.progress_callback(s)),
                result_callback=lambda res: self.root.after(0, lambda r=res: self._append_result_row(r))
            )
            results = self.bot_engine.process_credentials(file_path, parallel_browsers=parallel_browsers)
            
            # End session
            if results:
                success_count = sum(1 for r in results if r.get("status") == "success")
                invalid_count = sum(1 for r in results if r.get("status") == "invalid")
                banned_count = sum(1 for r in results if r.get("status") == "banned")
                error_count = sum(1 for r in results if r.get("status") == "error")
                
                self.db.end_session(
                    self.current_session_id,
                    len(results),
                    success_count,
                    invalid_count,
                    banned_count,
                    error_count
                )
            
            self.root.after(0, lambda: self.log("\n" + "=" * 60))
            self.root.after(0, lambda: self.log("Bot completed successfully!"))
            self.root.after(0, lambda: self.log("=" * 60))
            
            # Display results (update table in real-time)
            if results:
                self.root.after(0, lambda: self.display_results(results))
            
            # Reload stats
            self.root.after(0, self.load_stats)
            
            # Prepare resume file (remaining credentials) if any
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    tokens = [t.strip() for t in f.read().split() if t.strip()]
                processed = set()
                for r in results or []:
                    e = r.get("email") or ""
                    p = r.get("password") or ""
                    if e and p:
                        processed.add(f"{e}:{p}")
                remaining = [t for t in tokens if t not in processed]
                if remaining and len(remaining) < len(tokens):
                    resume_path = os.path.join(os.path.dirname(file_path), "credentials_resume.txt")
                    with open(resume_path, 'w', encoding='utf-8') as f:
                        f.write(" ".join(remaining))
                    # Switch Start button to Continue with resume file
                    self.root.after(0, lambda: self.file_path_var.set(resume_path))
                    self.root.after(0, lambda: self.start_btn.config(text="▶️  Continue", bg="#10b981", activebackground="#059669", state=tk.NORMAL))
                else:
                    self.root.after(0, lambda: self.start_btn.config(text="▶️  Start Bot", bg="#10b981", activebackground="#059669", state=tk.NORMAL))
            except Exception as ex:
                self.root.after(0, lambda: self.log(f"Resume preparation error: {str(ex)}"))
            
        except Exception as e:
            err_msg = f"\nError: {str(e)}"
            self.root.after(0, lambda msg=err_msg: self.log(msg))
        finally:
            self.root.after(0, self.stop_bot)

    
    def stop_bot(self):
        """Stop the bot and disconnect VPN."""
        self.is_running = False
        self._footer_running = False
        # Stop progress bar
        try:
            if hasattr(self, "batch_progress"):
                self.batch_progress.stop()
        except:
            pass
        # Stop bot engine (this will disconnect VPN)
        if self.bot_engine:
            self.bot_engine.stop()
        # Also disconnect GUI VPN manager if exists (always disconnect after stop)
        try:
            if self.vpn_manager:
                self.log("Disconnecting ExpressVPN...")
                self.vpn_manager.disconnect()
                # Force immediate status refresh
                self.update_vpn_status()
        except:
            pass
        # Reset button for next run
        self.start_btn.config(text="▶️  Start Bot", bg="#10b981", activebackground="#059669", state=tk.NORMAL)
        try:
            self.browse_btn.config(state=tk.NORMAL)
        except:
            pass
        try:
            self.parallel_spinbox.config(state="normal")
        except:
            pass
    
    def _create_stat_card(self, parent, label, value, color):
        """Create a compact stat card."""
        card = tk.Frame(parent, bg=color, relief=tk.FLAT, bd=0)
        card.pack(side=tk.LEFT, padx=4, fill=tk.BOTH, expand=True)
        
        content = tk.Frame(card, bg=color)
        content.pack(fill=tk.BOTH, expand=True, padx=8, pady=6)
        
        value_label = tk.Label(
            content,
            text=value,
            font=("Segoe UI", 14, "bold"),
            bg=color,
            fg="white"
        )
        value_label.pack()
        
        label_label = tk.Label(
            content,
            text=label,
            font=("Segoe UI", 8),
            bg=color,
            fg="white"
        )
        label_label.pack()
        
        return {"card": card, "value": value_label, "label": label_label}
    
    def log(self, message):
        """Add message to logs with syntax highlighting and filter support."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        
        # Determine log level and apply color
        message_lower = message.lower()
        if "✅" in message or "success" in message_lower:
            level = "success"
        elif "❌" in message or "error" in message_lower or "failed" in message_lower:
            level = "error"
        elif "⚠️" in message or "warning" in message_lower or "banned" in message_lower:
            level = "warning"
        else:
            level = "info"
        
        # Respect current log filter if set
        try:
            current_filter = self.log_filter.get()
        except Exception:
            current_filter = "all"
        if current_filter != "all" and level != current_filter:
            return
        
        # Insert timestamp with timestamp tag
        self.logs_text.insert(tk.END, f"[{timestamp}] ", "timestamp")
        # Insert message with appropriate tag
        self.logs_text.insert(tk.END, f"{message}\n", level)
        self.logs_text.see(tk.END)
        self.root.update_idletasks()
    
    def log_callback(self, message):
        """Callback for bot engine logs."""
        self.root.after(0, lambda: self.log(message))
    
    def display_results(self, results):
        """Display results in results tab with modern table view."""
        if not results:
            return
        
        # Store results for filtering
        self.all_results = results
        
        # Clear existing items
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)
        
        # Calculate summary
        success_count = sum(1 for r in results if r.get("status") == "success")
        invalid_count = sum(1 for r in results if r.get("status") == "invalid")
        banned_count = sum(1 for r in results if r.get("status") == "banned")
        error_count = sum(1 for r in results if r.get("status") == "error")
        
        # Update summary cards
        self.summary_total["value"].config(text=str(len(results)))
        self.summary_success["value"].config(text=str(success_count))
        self.summary_invalid["value"].config(text=str(invalid_count))
        self.summary_banned["value"].config(text=str(banned_count))
        self.summary_error["value"].config(text=str(error_count))
        
        # Add results to table
        for i, result in enumerate(results, 1):
            status = result.get("status", "unknown").lower()
            email = result.get("email", "N/A")
            username = result.get("username", "") or "N/A"
            karma = result.get("karma", "") or "N/A"
            ip = result.get("ip_address", "") or "N/A"
            country = result.get("country", "") or "N/A"
            location = result.get("location", "") or "N/A"
            reason = result.get("error_message", "") or "—"
            
            # Format status with icon
            status_icon = {
                "success": "✅ SUCCESS",
                "invalid": "❌ INVALID",
                "banned": "⚠️ BANNED",
                "error": "⚠️ ERROR"
            }.get(status, status.upper())
            
            # Insert row with appropriate tag for color coding
            item = self.results_tree.insert(
                "",
                tk.END,
                values=(
                    i,
                    email,
                    status_icon,
                    reason,
                    username,
                    karma,
                    ip,
                    country,
                    location
                ),
                tags=(status,)
            )
        
        # Auto-resize columns
        self._auto_resize_columns()

    def _append_result_row(self, result):
        """Append a single result row and update summaries live."""
        try:
            if not result:
                return
            # Keep in local results
            self.all_results.append(result)
            # Compute index
            i = len(self.all_results)
            status = (result.get("status") or "unknown").lower()
            email = result.get("email", "N/A")
            username = result.get("username", "") or "N/A"
            karma = result.get("karma", "") or "N/A"
            ip = result.get("ip_address", "") or "N/A"
            country = result.get("country", "") or "N/A"
            location = result.get("location", "") or "N/A"
            reason = result.get("error_message", "") or "—"
            status_icon = {
                "success": "✅ SUCCESS",
                "invalid": "❌ INVALID",
                "banned": "⚠️ BANNED",
                "error": "⚠️ ERROR"
            }.get(status, status.upper())
            self.results_tree.insert(
                "",
                tk.END,
                values=(i, email, status_icon, reason, username, karma, ip, country, location),
                tags=(status,)
            )
            # Update summaries
            success_count = sum(1 for r in self.all_results if (r.get("status") or "").lower() == "success")
            invalid_count = sum(1 for r in self.all_results if (r.get("status") or "").lower() == "invalid")
            banned_count = sum(1 for r in self.all_results if (r.get("status") or "").lower() == "banned")
            error_count = sum(1 for r in self.all_results if (r.get("status") or "").lower() == "error")
            self.summary_total["value"].config(text=str(len(self.all_results)))
            self.summary_success["value"].config(text=str(success_count))
            self.summary_invalid["value"].config(text=str(invalid_count))
            self.summary_banned["value"].config(text=str(banned_count))
            self.summary_error["value"].config(text=str(error_count))
        except:
            pass
    
    def _auto_resize_columns(self):
        """Auto-resize columns to fit content."""
        for col in self.results_tree["columns"]:
            self.results_tree.column(col, width=tk.font.Font().measure(col) + 20)
    
    def filter_results_table(self, *args):
        """Filter results table based on search term."""
        search_term = self.results_search_var.get().lower()
        
        # Clear existing items
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)
        
        # Filter and display results
        filtered_results = []
        if search_term:
            for result in self.all_results:
                # Search in email, username, status, country
                searchable_text = " ".join([
                    result.get("email", ""),
                    result.get("username", ""),
                    result.get("status", ""),
                    result.get("country", ""),
                    result.get("ip_address", "")
                ]).lower()
                
                if search_term in searchable_text:
                    filtered_results.append(result)
        else:
            filtered_results = self.all_results
        
        # Display filtered results
        if filtered_results:
            success_count = sum(1 for r in filtered_results if r.get("status") == "success")
            invalid_count = sum(1 for r in filtered_results if r.get("status") == "invalid")
            banned_count = sum(1 for r in filtered_results if r.get("status") == "banned")
            error_count = sum(1 for r in filtered_results if r.get("status") == "error")
            
            # Update summary cards
            self.summary_total["value"].config(text=str(len(filtered_results)))
            self.summary_success["value"].config(text=str(success_count))
            self.summary_invalid["value"].config(text=str(invalid_count))
            self.summary_banned["value"].config(text=str(banned_count))
            self.summary_error["value"].config(text=str(error_count))
            
            # Add filtered results to table
            for i, result in enumerate(filtered_results, 1):
                status = result.get("status", "unknown").lower()
                email = result.get("email", "N/A")
                username = result.get("username", "") or "N/A"
                karma = result.get("karma", "") or "N/A"
                ip = result.get("ip_address", "") or "N/A"
                country = result.get("country", "") or "N/A"
                location = result.get("location", "") or "N/A"
                reason = result.get("error_message", "") or "—"
                
                status_icon = {
                    "success": "✅ SUCCESS",
                    "invalid": "❌ INVALID",
                    "banned": "⚠️ BANNED",
                    "error": "⚠️ ERROR"
                }.get(status, status.upper())
                
                self.results_tree.insert(
                    "",
                    tk.END,
                    values=(
                        i,
                        email,
                        status_icon,
                        reason,
                        username,
                        karma,
                        ip,
                        country,
                        location
                    ),
                    tags=(status,)
                )
    
    def export_excel(self):
        """Export results to Excel."""
        if not self.bot_engine or not self.bot_engine.last_results:
            messagebox.showwarning("No Results", "No results to export. Run the bot first.")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if filename:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                from datetime import datetime
                
                results = self.bot_engine.last_results
                wb = Workbook()
                ws = wb.active
                ws.title = "Reddit Login Results"
                
                # Header row
                headers = ["Email", "Password", "Status", "Username", "Karma", "IP Address", "Country", "Location", "Error Message"]
                ws.append(headers)
                
                # Style header row
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                # Color coding
                status_colors = {
                    "success": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                    "invalid": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                    "banned": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
                    "error": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                }
                
                # Add data rows
                for result in results:
                    row = [
                        result.get("email", ""),
                        result.get("password", ""),
                        result.get("status", "").upper() if result.get("status") else "",
                        result.get("username", "") or "",
                        result.get("karma", "") or "",
                        result.get("ip_address", "") or "",
                        result.get("country", "") or "",
                        result.get("location", "") or "",
                        result.get("error_message", "") or ""
                    ]
                    ws.append(row)
                    
                    # Apply color to status column
                    status_cell = ws.cell(row=ws.max_row, column=3)
                    status = result.get("status", "").lower()
                    if status in status_colors:
                        status_cell.fill = status_colors[status]
                
                # Auto-adjust column widths
                column_widths = {"A": 30, "B": 25, "C": 12, "D": 20, "E": 15, "F": 18, "G": 20, "H": 30, "I": 50}
                for col, width in column_widths.items():
                    ws.column_dimensions[col].width = width
                
                # Wrap text for error message column
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=9, max_col=9):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                wb.save(filename)
                messagebox.showinfo("Success", f"Results exported to {filename}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to export: {str(e)}")
    
    def clear_results(self):
        """Clear results display."""
        # Clear table
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)
        
        # Clear stored results
        self.all_results = []
        
        # Reset summary cards
        self.summary_total["value"].config(text="0")
        self.summary_success["value"].config(text="0")
        self.summary_invalid["value"].config(text="0")
        self.summary_banned["value"].config(text="0")
        self.summary_error["value"].config(text="0")
        
        # Clear search
        self.results_search_var.set("")
    
    def logout(self):
        """Logout and return to login - clears saved credentials."""
        if self.is_running:
            if not messagebox.askyesno("Bot Running", "Bot is currently running. Stop and logout?"):
                return
            self.stop_bot()
        
        # Clear saved credentials on explicit logout
        try:
            saved_creds_file = "saved_credentials.json"
            if os.path.exists(saved_creds_file):
                os.remove(saved_creds_file)
        except:
            pass
        
        # End session if exists
        if self.current_session_id:
            try:
                self.db.end_session(self.current_session_id)
            except:
                pass
        
        self.root.destroy()
        # Show login window again
        login_root = tk.Tk()
        LoginWindow(login_root)
        login_root.mainloop()


def main():
    """Main entry point."""
    root = tk.Tk()
    app = LoginWindow(root)
    root.mainloop()


if __name__ == "__main__":
    main()

